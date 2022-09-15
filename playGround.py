# name = "file.txt"
# with open(name, 'w') as a:
#     addToFile = a.w

from multiprocessing.heap import Arena
from re import A
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook('authenticate_user/example.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet1'] # Get a sheet from the workbook.
# print(type(wb))
# print(type(sheet))
# print(sheet.title)
# anotherSheet = wb.active
# print(anotherSheet)
# print(sheet["A1"].value)
c = sheet["B1"] # Get a cell from the sheet.
# print(type(c.value)) # Get the value from the cell.
# a1Value = sheet['A1']
# print(type(a1Value))

# Get the row, column, and value from the cell

# x = 'row %s, column %s is %s' % (c.row, c.column, c.value)
x = "Row {}, Column {} is {}'".format(c.row, c.column, c.value)
print(x)
y = "Cell {} is {}'".format(c.coordinate, c.value)
print(y)

print(sheet['C1'].value)

print(sheet.cell(1,2))
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
# for i in range(1,8):
#     print(i,sheet.cell(i,3).value)

print(get_column_letter(1)) # Translate column 1 to a letter
print(column_index_from_string('A')) # Get A's number.
print(sheet.max_column) # Get the highest column number.
print(sheet.max_row) # Get the highest row number.
n = "73"
myList = []
for i in range(1,sheet.max_row+1):
    a = sheet.cell(i,2).value
    myList.append(a)

print(myList.index("aya12"))
