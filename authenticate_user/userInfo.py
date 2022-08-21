# import openpyxl
import openpyxl
from openpyxl import Workbook, load_workbook
def userNames():
    
    return userNames
# print(userNames())

def userPasswords():
    wb = load_workbook('authenticate_user/example.xlsx')
    ws = wb['Sheet1']
    passwords = []
    for i in range(2, ws.max_row+1):
        passwrd = ws["B{}".format(i)].value
        passwords.append(passwrd)
    return passwords

# print(userPasswords())

# def run():
#     passwords = userPasswords()
#     user_names = userNames()
#     return user_names, passwords

# print(run())